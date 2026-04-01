# ============================================================
# retry.py — Retraitement des lignes en erreur
# ============================================================
# Lit output_final.xlsx, repère les lignes avec un statut d'erreur,
# les supprime de output_final.xlsx et de progress.json,
# puis invite à relancer python main.py.
#
# Statuts retraités :
#   erreur, accès_refusé, profil_sans_exp
#
# IMPORTANT : ne modifie jamais gema.xlsx.
#
# Usage :
#   source .venv/bin/activate
#   python retry.py
# ============================================================

import os
from collections import Counter
from pathlib import Path

import openpyxl

import config
from utils.logger  import get_logger
from utils.helpers import charger_progression, sauvegarder_progression

log = get_logger("RETRY")

# Statuts qui méritent d'être retraités
STATUTS_A_RETENTER = {"erreur", "accès_refusé", "profil_sans_exp", "profil_non_trouvé"}


# ----------------------------------------------------------
# Lecture des erreurs dans output_final.xlsx
# ----------------------------------------------------------

def charger_erreurs(chemin: str) -> list:
    """
    Lit output_final.xlsx et retourne la liste des dicts
    pour les lignes dont le statut est à retenter.

    Chaque dict contient : id, nom, prenom, statut_precedent.
    """
    if not Path(chemin).exists():
        log.error(f"Fichier introuvable : {chemin}")
        return []

    wb = openpyxl.load_workbook(chemin)
    ws = wb.active

    # Mapping entête → index (0-based)
    entetes = {str(c.value or "").strip(): i for i, c in enumerate(ws[1])}

    idx_id     = entetes.get("id", -1)
    idx_nom    = entetes.get("Nom", -1)
    idx_prenom = entetes.get("Prénom", -1)
    idx_statut = entetes.get("Statut", -1)

    if idx_statut < 0:
        log.error("Colonne 'Statut' introuvable dans output_final.xlsx.")
        return []

    erreurs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def g(i):
            return str(row[i] if i >= 0 and i < len(row) and row[i] is not None else "").strip()

        statut = g(idx_statut)
        if statut not in STATUTS_A_RETENTER:
            continue

        nom    = g(idx_nom)
        prenom = g(idx_prenom)
        if not nom and not prenom:
            continue

        erreurs.append({
            "id":               g(idx_id),
            "nom":              nom,
            "prenom":           prenom,
            "statut_precedent": statut,
        })

    return erreurs


# ----------------------------------------------------------
# Nettoyage de output_final.xlsx
# ----------------------------------------------------------

def supprimer_lignes_erreur(chemin: str, cles: set) -> int:
    """
    Supprime les lignes correspondant aux clés (id ou nom_prenom) dans output_final.xlsx.
    Retourne le nombre de lignes supprimées.
    Sauvegarde atomique.
    """
    wb = openpyxl.load_workbook(chemin)
    ws = wb.active

    entetes    = {str(c.value or "").strip(): i for i, c in enumerate(ws[1])}
    idx_id     = entetes.get("id", -1)
    idx_nom    = entetes.get("Nom", -1)
    idx_prenom = entetes.get("Prénom", -1)

    lignes_a_supprimer = []
    for row in ws.iter_rows(min_row=2):
        id_cell     = str(row[idx_id].value     or "").strip() if idx_id     >= 0 else ""
        nom_cell    = str(row[idx_nom].value    or "").strip() if idx_nom    >= 0 else ""
        prenom_cell = str(row[idx_prenom].value or "").strip() if idx_prenom >= 0 else ""

        cle = id_cell if id_cell else f"{nom_cell}_{prenom_cell}"
        if cle in cles:
            lignes_a_supprimer.append(row[0].row)

    # Supprimer en ordre inverse pour ne pas décaler les indices
    for num in sorted(lignes_a_supprimer, reverse=True):
        ws.delete_rows(num)

    tmp = chemin + ".tmp"
    wb.save(tmp)
    os.replace(tmp, chemin)

    return len(lignes_a_supprimer)


# ----------------------------------------------------------
# Point d'entrée
# ----------------------------------------------------------

def main():
    log.info("=" * 60)
    log.info("  retry.py — Retraitement des erreurs")
    log.info("=" * 60)

    # 1. Identifier les erreurs
    erreurs = charger_erreurs(config.FICHIER_OUTPUT_FINAL)
    if not erreurs:
        log.info("Aucune erreur à retraiter. Terminé.")
        return

    # Afficher un résumé par statut
    comptages = Counter(e["statut_precedent"] for e in erreurs)
    log.info(f"{len(erreurs)} ligne(s) à retraiter :")
    for statut, nb in comptages.most_common():
        log.info(f"  {statut:<20} : {nb}")

    print()
    for e in erreurs[:20]:
        log.info(f"  - {e['prenom']} {e['nom']}  ({e['statut_precedent']})")
    if len(erreurs) > 20:
        log.info(f"  ... et {len(erreurs) - 20} autre(s)")

    # 2. Confirmation
    print()
    try:
        reponse = input(
            f"  Supprimer ces {len(erreurs)} ligne(s) et les remettre en file ? (o/n) : "
        ).strip().lower()
    except (KeyboardInterrupt, EOFError):
        reponse = "n"

    if reponse != "o":
        log.info("Annulé.")
        return

    # Construire le set de clés (id si disponible, sinon nom_prenom)
    cles = {
        e["id"] if e["id"] else f"{e['nom']}_{e['prenom']}"
        for e in erreurs
    }

    # 3. Supprimer les lignes de output_final.xlsx
    nb_suppr = supprimer_lignes_erreur(config.FICHIER_OUTPUT_FINAL, cles)
    log.info(f"{nb_suppr} ligne(s) supprimée(s) de {config.FICHIER_OUTPUT_FINAL}.")

    # 4. Retirer ces personnes de progress.json
    progression = charger_progression(config.FICHIER_PROGRESSION)
    traites     = set(progression.get("traites", []))

    avant = len(traites)
    traites -= cles
    progression["traites"] = list(traites)
    sauvegarder_progression(config.FICHIER_PROGRESSION, progression)
    log.info(f"{avant - len(traites)} clé(s) retirée(s) de progress.json.")

    # 5. Instruction finale
    log.info("")
    log.info("Relancez maintenant :  python main.py")
    log.info("Le script retraitera automatiquement ces personnes.")


if __name__ == "__main__":
    main()
