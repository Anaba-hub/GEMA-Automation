# ============================================================
# agents/agent_excel.py — Lecture et écriture des fichiers Excel
# ============================================================
# Source unique : gema.xlsx
# Colonnes gema : id | Prénom | Nom | Moyenne /4 | Année | École |
#                 Recherche LinkedIn | Poste 1 | Poste 2 | Vérifié GEMA
#
# Sortie : output_final.xlsx (14 colonnes)
# Colonnes sortie : id | Prénom | Nom | Année | École |
#                   Lien retenu | Lien source |
#                   Poste 1 | Société 1 | Période 1 |
#                   Poste 2 | Société 2 | Période 2 | Statut
#
# Lien retenu  = URL /in/ utilisée pour l'extraction
# Lien source  = URL d'origine dans gema (peut être /search/ ou /in/)
# ============================================================

import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from utils.logger import get_logger

log = get_logger("EXCEL")

# En-têtes du fichier de sortie (14 colonnes)
ENTETES_OUTPUT = [
    "id", "Prénom", "Nom", "Année", "École",
    "Lien retenu", "Lien source",
    "Poste 1", "Société 1", "Période 1",
    "Poste 2", "Société 2", "Période 2",
    "Statut",
]

# Valeurs considérées comme vides/invalides dans les cellules
_VIDES = {"", "nan", "#n/a", "#n/a!", "non trouvé", "none", "n/a", "-", "#value!"}


# ----------------------------------------------------------
# Fonctions publiques
# ----------------------------------------------------------

def est_lien_valide(valeur) -> bool:
    """
    Retourne True si la valeur est une URL LinkedIn utilisable.
    Rejette : None, "", "nan", "#N/A", "Non trouvé", toute chaîne
    sans "linkedin.com" ou ne commençant pas par "http".
    """
    if not valeur:
        return False
    s = str(valeur).strip()
    if s.lower() in _VIDES:
        return False
    if not s.startswith("http"):
        return False
    if "linkedin.com" not in s.lower():
        return False
    return True


def load_gema(chemin: str = "gema.xlsx") -> list:
    """
    Lit gema.xlsx et retourne une liste de dicts.

    Chaque dict contient :
      id, prenom, nom, annee, ecole, lien,
      poste1_gema, poste2_gema, deja_traite

    deja_traite = True si Poste 1 ET Poste 2 sont déjà remplis dans gema.
    Gère les cellules vides, "nan", "#N/A" proprement.
    """
    if not Path(chemin).exists():
        log.error(f"Fichier introuvable : {chemin}")
        return []

    wb = openpyxl.load_workbook(chemin)
    ws = wb.active

    # Mapping entête → index colonne (0-based)
    entetes = {}
    for idx, cell in enumerate(ws[1]):
        if cell.value:
            entetes[str(cell.value).strip()] = idx

    def val(row, key):
        """Valeur texte d'une colonne, ou '' si absente/invalide."""
        idx = entetes.get(key, -1)
        if idx < 0 or idx >= len(row):
            return ""
        v = row[idx].value
        if v is None:
            return ""
        s = str(v).strip()
        return "" if s.lower() in _VIDES else s

    personnes = []
    for row in ws.iter_rows(min_row=2):
        nom    = val(row, "Nom")
        prenom = val(row, "Prénom")
        if not nom and not prenom:
            continue  # ligne vide

        poste1 = val(row, "Poste 1")
        poste2 = val(row, "Poste 2")

        personnes.append({
            "id":          val(row, "id"),
            "prenom":      prenom,
            "nom":         nom,
            "annee":       val(row, "Année"),
            "ecole":       val(row, "École"),
            "lien":        val(row, "Recherche LinkedIn"),
            "poste1_gema": poste1,
            "poste2_gema": poste2,
            # deja_traite : les deux postes sont déjà connus dans gema → skip extraction
            "deja_traite": bool(poste1 and poste2),
        })

    log.info(f"{len(personnes)} personne(s) chargée(s) depuis '{chemin}'.")
    return personnes


# ----------------------------------------------------------
# OutputWriter — gestion optimisée du fichier de sortie
# ----------------------------------------------------------

class OutputWriter:
    """
    Gestionnaire du fichier output_final.xlsx.

    Garde le workbook ouvert en mémoire pour éviter de relire/réécrire
    le fichier à chaque ligne. Sauvegarde périodiquement (toutes les
    N lignes) et à la fermeture.
    """

    def __init__(self, chemin: str, intervalle_sauvegarde: int = 10):
        self.chemin = chemin
        self.intervalle = intervalle_sauvegarde
        self._compteur_depuis_sauvegarde = 0
        self._wb = None
        self._ws = None
        self._idx_map = {}

    def ouvrir(self) -> None:
        """Ouvre ou crée le fichier output_final.xlsx."""
        if not Path(self.chemin).exists():
            self._creer_output()

        self._wb = openpyxl.load_workbook(self.chemin)
        self._ws = self._wb.active
        self._idx_map = {
            str(cell.value or "").strip(): cell.column
            for cell in self._ws[1]
        }
        log.info(f"OutputWriter ouvert : {self.chemin}")

    def save_row(self, data: dict) -> None:
        """
        Ajoute ou met à jour une ligne dans le workbook en mémoire.
        Sauvegarde sur disque toutes les N lignes.
        """
        if not self._ws:
            raise RuntimeError("OutputWriter non ouvert. Appelez ouvrir() d'abord.")

        id_data     = str(data.get("id", "")).strip()
        nom_data    = str(data.get("nom", "")).strip()
        prenom_data = str(data.get("prenom", "")).strip()

        # Chercher une ligne existante à mettre à jour
        ligne_cible = None
        col_id     = self._idx_map.get("id", 1)
        col_nom    = self._idx_map.get("Nom", 3)
        col_prenom = self._idx_map.get("Prénom", 2)

        for row in self._ws.iter_rows(min_row=2):
            id_cell     = str(row[col_id - 1].value or "").strip()
            nom_cell    = str(row[col_nom - 1].value or "").strip()
            prenom_cell = str(row[col_prenom - 1].value or "").strip()

            if id_data and id_cell == id_data:
                ligne_cible = row[0].row
                break
            if not id_data and nom_cell == nom_data and prenom_cell == prenom_data:
                ligne_cible = row[0].row
                break

        # Valeurs à écrire dans l'ordre de ENTETES_OUTPUT
        valeurs = [
            data.get("id", ""),
            data.get("prenom", ""),
            data.get("nom", ""),
            data.get("annee", ""),
            data.get("ecole", ""),
            data.get("lien_retenu", ""),
            data.get("lien_source", ""),
            data.get("poste1", ""),
            data.get("societe1", ""),
            data.get("periode1", ""),
            data.get("poste2", ""),
            data.get("societe2", ""),
            data.get("periode2", ""),
            data.get("statut", ""),
        ]

        if ligne_cible:
            for col_idx, valeur in enumerate(valeurs, start=1):
                self._ws.cell(row=ligne_cible, column=col_idx, value=valeur)
            log.info(f"Mise à jour : {prenom_data} {nom_data} → {data.get('statut', '')}")
        else:
            self._ws.append(valeurs)
            log.info(f"Nouvelle ligne : {prenom_data} {nom_data} → {data.get('statut', '')}")

        self._compteur_depuis_sauvegarde += 1
        if self._compteur_depuis_sauvegarde >= self.intervalle:
            self._sauvegarder_disque()

    def fermer(self) -> None:
        """Ajuste les largeurs de colonnes, sauvegarde et ferme le workbook."""
        if not self._wb:
            return

        # Ajustement des largeurs uniquement à la fermeture
        self._ajuster_largeurs()
        self._sauvegarder_disque()
        self._wb = None
        self._ws = None
        log.info(f"OutputWriter fermé : {self.chemin}")

    def _sauvegarder_disque(self) -> None:
        """Sauvegarde atomique sur disque."""
        if not self._wb:
            return
        chemin_tmp = self.chemin + ".tmp"
        self._wb.save(chemin_tmp)
        os.replace(chemin_tmp, self.chemin)
        self._compteur_depuis_sauvegarde = 0
        log.debug(f"Sauvegarde disque : {self.chemin}")

    def _ajuster_largeurs(self) -> None:
        """Ajuste la largeur des colonnes (appelé uniquement à la fermeture)."""
        if not self._ws:
            return
        for col in self._ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            self._ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 70)

    def _creer_output(self) -> None:
        """Crée output_final.xlsx avec les en-têtes stylisés."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Résultats"
        ws.append(ENTETES_OUTPUT)

        fill  = PatternFill(fill_type="solid", fgColor="1F4E79")
        font  = Font(bold=True, color="FFFFFF", size=11)
        align = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.fill      = fill
            cell.font      = font
            cell.alignment = align
        ws.row_dimensions[1].height = 20

        wb.save(self.chemin)
        log.info(f"Fichier créé : {self.chemin}")


# ----------------------------------------------------------
# Fonction legacy (utilisée par retry.py)
# ----------------------------------------------------------

def save_row(chemin: str, data: dict) -> None:
    """
    Sauvegarde une ligne dans output_final.xlsx (mode standalone).
    Utilisé par retry.py qui ne maintient pas un OutputWriter ouvert.
    """
    writer = OutputWriter(chemin)
    writer.ouvrir()
    writer.save_row(data)
    writer.fermer()
