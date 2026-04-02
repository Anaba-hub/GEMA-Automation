import json
import os
from datetime import datetime
from pathlib import Path

import openpyxl


# ----------------------------------------------------------
# Progression (progress.json)
# ----------------------------------------------------------

def charger_progression(chemin):
    path = Path(chemin)
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            import shutil
            shutil.copy(chemin, chemin + ".backup")
            return {}
    return {}


def sauvegarder_progression(chemin, progression):
    progression["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    chemin_tmp = chemin + ".tmp"
    with open(chemin_tmp, "w", encoding="utf-8") as f:
        json.dump(progression, f, ensure_ascii=False, indent=2)
    os.replace(chemin_tmp, chemin)


# ----------------------------------------------------------
# Barre de progression
# ----------------------------------------------------------

def afficher_barre(actuel, total, largeur=40, suffixe=""):
    pct    = actuel / total if total else 0
    rempli = int(largeur * pct)
    barre  = "\u2588" * rempli + "\u2591" * (largeur - rempli)
    print(f"\r  [{barre}] {actuel}/{total} ({pct*100:.1f}%) {suffixe}", end="", flush=True)
    if actuel == total:
        print()


# ----------------------------------------------------------
# Retry : nettoyage des erreurs pour relance
# ----------------------------------------------------------

STATUTS_A_RETENTER = {"erreur", "acces_refuse", "profil_sans_exp", "profil_non_trouve",
                      "accès_refusé", "profil_non_trouvé"}


def nettoyer_erreurs(chemin_output, chemin_progression):
    """
    Supprime les lignes en erreur de output_final.xlsx et progress.json.
    Retourne le nombre de lignes supprimees.
    """
    if not Path(chemin_output).exists():
        return 0

    wb = openpyxl.load_workbook(chemin_output)
    ws = wb.active

    entetes = {str(c.value or "").strip(): i for i, c in enumerate(ws[1])}
    idx_statut = entetes.get("Statut", -1)
    idx_id     = entetes.get("id", -1)
    idx_nom    = entetes.get("Nom", -1)
    idx_prenom = entetes.get("Prénom", -1)
    idx_annee  = entetes.get("Année", -1)
    idx_ecole  = entetes.get("École", -1)

    if idx_statut < 0:
        return 0

    lignes_a_supprimer = []
    cles_a_retirer = set()

    for row in ws.iter_rows(min_row=2):
        statut = str(row[idx_statut].value or "").strip()
        if statut not in STATUTS_A_RETENTER:
            continue

        lignes_a_supprimer.append(row[0].row)

        # Construire la cle identique a main.py
        id_val     = str(row[idx_id].value or "").strip() if idx_id >= 0 else ""
        nom_val    = str(row[idx_nom].value or "").strip() if idx_nom >= 0 else ""
        prenom_val = str(row[idx_prenom].value or "").strip() if idx_prenom >= 0 else ""
        annee_val  = str(row[idx_annee].value or "").strip() if idx_annee >= 0 else ""
        ecole_val  = str(row[idx_ecole].value or "").strip() if idx_ecole >= 0 else ""

        cle = id_val if id_val else f"{nom_val}_{prenom_val}_{annee_val}_{ecole_val}"
        cles_a_retirer.add(cle)

    # Supprimer les lignes (ordre inverse)
    for num in sorted(lignes_a_supprimer, reverse=True):
        ws.delete_rows(num)

    tmp = chemin_output + ".tmp"
    wb.save(tmp)
    os.replace(tmp, chemin_output)

    # Nettoyer progress.json
    progression = charger_progression(chemin_progression)
    traites = set(progression.get("traites", []))
    traites -= cles_a_retirer
    progression["traites"] = list(traites)
    sauvegarder_progression(chemin_progression, progression)

    return len(lignes_a_supprimer)
