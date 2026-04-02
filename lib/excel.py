import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from lib.logger import get_logger

log = get_logger("EXCEL")

ENTETES_OUTPUT = [
    "id", "Prénom", "Nom", "Année", "École",
    "Lien retenu", "Lien source",
    "Poste 1", "Société 1", "Période 1",
    "Poste 2", "Société 2", "Période 2",
    "Statut",
]

_VIDES = {"", "nan", "#n/a", "#n/a!", "non trouvé", "none", "n/a", "-", "#value!"}


def est_lien_valide(valeur):
    if not valeur:
        return False
    s = str(valeur).strip()
    if s.lower() in _VIDES:
        return False
    return s.startswith("http") and "linkedin.com" in s.lower()


def load_gema(chemin="gema.xlsx"):
    if not Path(chemin).exists():
        log.error(f"Fichier introuvable : {chemin}")
        return []

    wb = openpyxl.load_workbook(chemin)
    ws = wb.active

    entetes = {}
    for idx, cell in enumerate(ws[1]):
        if cell.value:
            entetes[str(cell.value).strip()] = idx

    def val(row, key):
        idx = entetes.get(key, -1)
        if idx < 0 or idx >= len(row):
            return ""
        v = row[idx].value
        if v is None:
            return ""
        s = str(v).strip()
        return "" if s.lower() in _VIDES else s

    personnes = []
    for row in ws.iter_rows(min_row=2):
        nom    = val(row, "Nom")
        prenom = val(row, "Prénom")
        if not nom and not prenom:
            continue

        poste1 = val(row, "Poste 1")
        poste2 = val(row, "Poste 2")

        personnes.append({
            "id":          val(row, "id"),
            "prenom":      prenom,
            "nom":         nom,
            "annee":       val(row, "Année"),
            "ecole":       val(row, "École"),
            "lien":        val(row, "Recherche LinkedIn"),
            "poste1_gema": poste1,
            "poste2_gema": poste2,
            "deja_traite": bool(poste1 and poste2),
        })

    log.info(f"{len(personnes)} personne(s) chargée(s) depuis '{chemin}'.")
    return personnes


class OutputWriter:

    def __init__(self, chemin, intervalle_sauvegarde=10):
        self.chemin = chemin
        self.intervalle = intervalle_sauvegarde
        self._compteur = 0
        self._wb = None
        self._ws = None
        self._idx_map = {}

    def ouvrir(self):
        if not Path(self.chemin).exists():
            self._creer_output()
        self._wb = openpyxl.load_workbook(self.chemin)
        self._ws = self._wb.active
        self._idx_map = {
            str(cell.value or "").strip(): cell.column
            for cell in self._ws[1]
        }

    def save_row(self, data):
        if not self._ws:
            raise RuntimeError("OutputWriter non ouvert.")

        id_data     = str(data.get("id", "")).strip()
        nom_data    = str(data.get("nom", "")).strip()
        prenom_data = str(data.get("prenom", "")).strip()

        ligne_cible = None
        col_id     = self._idx_map.get("id", 1)
        col_nom    = self._idx_map.get("Nom", 3)
        col_prenom = self._idx_map.get("Prénom", 2)

        for row in self._ws.iter_rows(min_row=2):
            id_cell     = str(row[col_id - 1].value or "").strip()
            nom_cell    = str(row[col_nom - 1].value or "").strip()
            prenom_cell = str(row[col_prenom - 1].value or "").strip()

            if id_data and id_cell == id_data:
                ligne_cible = row[0].row
                break
            if not id_data and nom_cell == nom_data and prenom_cell == prenom_data:
                ligne_cible = row[0].row
                break

        valeurs = [
            data.get("id", ""), data.get("prenom", ""), data.get("nom", ""),
            data.get("annee", ""), data.get("ecole", ""),
            data.get("lien_retenu", ""), data.get("lien_source", ""),
            data.get("poste1", ""), data.get("societe1", ""), data.get("periode1", ""),
            data.get("poste2", ""), data.get("societe2", ""), data.get("periode2", ""),
            data.get("statut", ""),
        ]

        if ligne_cible:
            for col_idx, valeur in enumerate(valeurs, start=1):
                self._ws.cell(row=ligne_cible, column=col_idx, value=valeur)
        else:
            self._ws.append(valeurs)

        self._compteur += 1
        if self._compteur >= self.intervalle:
            self._sauvegarder_disque()

    def fermer(self):
        if not self._wb:
            return
        self._ajuster_largeurs()
        self._sauvegarder_disque()
        self._wb = None
        self._ws = None

    def _sauvegarder_disque(self):
        if not self._wb:
            return
        tmp = self.chemin + ".tmp"
        self._wb.save(tmp)
        os.replace(tmp, self.chemin)
        self._compteur = 0

    def _ajuster_largeurs(self):
        if not self._ws:
            return
        for col in self._ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            self._ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 70)

    def _creer_output(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Résultats"
        ws.append(ENTETES_OUTPUT)

        fill  = PatternFill(fill_type="solid", fgColor="1F4E79")
        font  = Font(bold=True, color="FFFFFF", size=11)
        align = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.fill      = fill
            cell.font      = font
            cell.alignment = align

        wb.save(self.chemin)
